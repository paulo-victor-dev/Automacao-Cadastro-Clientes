import requests
import json
import openpyxl
import re
import unicodedata
import os
import sys
import pyautogui as pg
import pyperclip
from time import sleep, time
from datetime import datetime, timedelta
import win32com.client as win32
import subprocess
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver import ActionChains


class Request_API:
    def _consultar_cnpj(self, cnpj):
        try:
            response = requests.get(f'https://open.cnpja.com/office/{cnpj}')
            resp = json.loads(response.text)
            if resp.get('code') == 400:
                print(f"{cnpj}:\n - ERRO AO CONSULTAR O CNPJ")
                return None
            else:
                return resp
        except Exception as e:
            print(f"{cnpj}:\n - ERRO AO CONSULTAR O CNPJ\n{e}")
            return None
            
    def _verif_situacao_ramo(self, resp):
        situação = resp.get('status', {}).get('text', '')
        situacao_se_ativa = True if situação == 'Ativa' else False

        atv_principal = resp.get('mainActivity', {}).get('text', '')
        atvs_secundarias = resp.get('sideActivities', [])
        lista_atvs = []

        palavras_atv_principal = atv_principal.replace(',', '').replace('.', '').split(' ')
        lista_atvs.extend(palavras_atv_principal)

        lista_atvs.extend(atvs_sec for atv in atvs_secundarias for atvs_sec in atv['text'].replace(',', '').replace('.', '').split(' '))
        
        for palavra in lista_atvs:
            if palavra.lower() in ['vestuário', 'vestuario', 'armarinho', 'magazine', 'magazines']:
                se_ramo = True
                break
            else:
                se_ramo = False

        return situacao_se_ativa, se_ramo
    
    def _verif_inscricao(self, resp):
        inscricao_est = resp['registrations']

        if inscricao_est:
            se_habilitada = resp['registrations'][0].get('enabled')
            num_inscricao = resp['registrations'][0].get('number') if se_habilitada else None
        else:
            num_inscricao = None

        return num_inscricao
    
    def _verif_razao(self, resp):
        razao = resp.get('company', {}).get('name', '').upper()
        sep_razao = razao.split(' ')

        remover = ['DA','DE','DI','DO','DU','DAS','DOS','E','&','-',"'",'.',',']

        lista_razao = [palavra for pos, palavra in enumerate(sep_razao) if not re.search(r'\d+', sep_razao[pos]) and palavra not in remover]

        lista_razao_ajustada = [palavra for pos, palavra in enumerate(lista_razao) if len(palavra) >= 3]

        if len(lista_razao_ajustada) > 2:
            razao_final = lista_razao_ajustada[0:2]
        else:
            razao_final = lista_razao_ajustada[0:]

        razao_final = ' '.join(razao_final)

        return razao_final

    def _extrair_endereco(self, resp):
        bairro = resp.get('address', {}).get('district', '').upper()
        cidade = resp.get('address', {}).get('city', '').upper()
        uf = resp.get('address', {}).get('state', '').upper()

        return bairro, cidade, uf

    def _contar_execucoes(self, contador, tempo_inicio):
        formatacoes = Formatacoes()

        tempo_maximo = 60

        if contador == 0:
            tempo_inicio = time()

        if contador == 5:
            tempo_execucao = time() - tempo_inicio
            if tempo_execucao < tempo_maximo:
                tempo_espera = tempo_maximo - tempo_execucao

                formatacoes._formatar_titulos(f'Aguardando {tempo_espera:.2f} para continuar...')
                sleep(tempo_espera)

            contador = 0
            tempo_inicio = 0
        else:
            contador += 1

        return contador, tempo_inicio


class Formatacoes:
    def _formatar_cnpj_consulta(self, cnpj):
        cnpj_formatado = cnpj.replace('.', '').replace('/', '').replace('-', '').replace(' ', '').strip()

        return cnpj_formatado

    def _formatar_cnpj_original(self, cnpj):
        cnpj_formatado = f'{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}'

        return cnpj_formatado

    def _formatar_titulos(self, titulo):
        print('=' * 60)
        print(f'{titulo}'.center(60).upper())
        print('=' * 60)

    def _remover_acentos(self, palavra):
        plvr = unicodedata.normalize('NFD', palavra)

        plvr_sem_acento = ''.join(c for c in plvr if unicodedata.category(c) != 'Mn')

        return plvr_sem_acento


class Planilhas:
    def __init__(self):
        self.caminho_plan_clientes = r'c:\Users\Cliente\OneDrive\Desktop\Verificar Cadastros\Clientes_para_cadastrar_V2.xlsx'
        self.caminho_plan_manse = r'c:\Users\Cliente\OneDrive\Desktop\Verificar Cadastros\Base de Clientes Manse.xlsx'
        self.formatacoes = Formatacoes()

    def _carregar_planilha_clientes(self):
        self.planilha_clientes = openpyxl.load_workbook(self.caminho_plan_clientes)
        planilha_ativa = self.planilha_clientes.active

        return planilha_ativa
    
    def _carregar_planilha_manse(self):
        self.planilha_manse = openpyxl.load_workbook(self.caminho_plan_manse)
        planilha_ativa = self.planilha_manse.active

        return planilha_ativa

    def _salvar_planilha(self, planilha, caminho_orignal):
        caminho_temporario = caminho_orignal + '.temp'

        try:
            planilha.save(caminho_temporario)
            os.replace(caminho_temporario, caminho_orignal)
            print(' - PLANILHA "CLIENTES PARA CADASTRAR" SALVA COM SUCESSO!')
            print('=' * 60)

        except Exception as e:
            print(f'ERRO AO SALVAR A PLANILHA: {e}')
            if os.path.exists(caminho_temporario):
                os.remove(caminho_temporario)
            sys.exit(1)

    def _verif_praca(self, cnpj_cliente, bairro_cnpj, cidade_cnpj, uf_cnpj):
        plan_manse = self._carregar_planilha_manse()
        datas = []

        cnpj_cliente = self.formatacoes._formatar_cnpj_original(cnpj_cliente)
        bairro_cnpj = self.formatacoes._remover_acentos(bairro_cnpj)
        cidade_cnpj = self.formatacoes._remover_acentos(cidade_cnpj)

        if cidade_cnpj in ['REDENCAO', 'IMPERATRIZ', 'OEIRAS', 'FOZ DO IGUACU', 'MACAPA', 'BOA VISTA' if uf_cnpj == 'RR' else '', 'SANTOS' if uf_cnpj == 'SP' else '']:
            print(' - SEM PRAÇA')
            return False

        for linha in plan_manse.iter_rows(min_row=2):
            cnpj_manse = linha[3].value
            bairro_manse = linha[10].value.upper()
            cidade_manse = linha[11].value.upper()
            uf_manse = linha[12].value

            if bairro_cnpj == bairro_manse and cidade_cnpj == cidade_manse and uf_cnpj == uf_manse and cnpj_manse != cnpj_cliente:
                dt_ult_compra = linha[9].value

                try:
                    data_formatada = datetime.strptime(dt_ult_compra, "%d/%m/%Y")
                    datas.append(data_formatada)
                except Exception as e:
                    print(f'Erro em formatar a data:\n{e}')
                    sys.exit(1)
        
        maior_dt = max(datas) if datas else None
        dif_data = datetime.now() - maior_dt if maior_dt else None

        if dif_data and timedelta(days=0) <= dif_data <= timedelta(days=30):
            print(' - SEM PRAÇA')
            return False
        else:
            return True

    def _verif_ja_cadastrado(self, cnpj_cliente):
        plan_manse = self._carregar_planilha_manse()

        for linha in plan_manse.iter_rows(min_row=2):
            cnpj_manse = linha[3].value

            cnpj_manse_format = self.formatacoes._formatar_cnpj_consulta(cnpj_manse)

            if cnpj_manse_format == cnpj_cliente:
                dt_cadastro = linha[0].value
                nome_vendedora = linha[7].value

                print(' - CLIENTE JÁ CADASTRADO')

                return dt_cadastro, nome_vendedora, True
        
        return None, None, False

    def _att_plan_manse(self):
        tentativas = 0
        max_tentativas = 3

        while tentativas < max_tentativas:
            try:
                excel = win32.Dispatch('Excel.Application')

                planilha = excel.Workbooks.Open(self.caminho_plan_manse)
                sleep(1)

                planilha.RefreshAll()
                sleep(2.5)

                planilha.Close(SaveChanges=True)
                sleep(1)

                excel.Quit()
                break

            except Exception as e:
                tentativas += 1
                print('Erro em atualizar PLANILHA MANSE. Tentando novamente...')
                sleep(1.5)

                try:
                    excel.Quit()
                except:
                    pass

                if tentativas == max_tentativas:
                    print(f'Erro em atualizar PLANILHA MANSE: Tentativas excedidas!')
                    sys.exit(1)


class Vendedoras:
    def __init__(self):
        self.lista_vendedoras = ['SIMONE DJ', 'LUCIANA', 'VIVIANE', 'JULIANA']
        self.caminho_arq_log = r'Log_nomes_vendedoras.txt'
        self.dic_vendedoras = [
            {
                'nome': 'SIMONE DJ',
                'img_cad_manse': r'Imagens_manse_vendedoras\SimoneDJ.png',
                'resp_rapida_cad_whats': '/SimoneDJ',
                'resp_rapida_ja_cad_whats': '/ClienteJaCadastradoSimoDJ'
            },

            {
                'nome': 'LUCIANA',
                'img_cad_manse': r'Imagens_manse_vendedoras\Luciane.png',
                'resp_rapida_cad_whats': '/Luciane',
                'resp_rapida_ja_cad_whats': '/ClienteJaCadastradoLu'
            },

            {
                'nome': 'VIVIANE',
                'img_cad_manse': r'Imagens_manse_vendedoras\Viviane.png',
                'resp_rapida_cad_whats': '/Viviane',
                'resp_rapida_ja_cad_whats': '/ClienteJaCadastradoVivi'
            },

            {
                'nome': 'JULIANA',
                'img_cad_manse': r'Imagens_manse_vendedoras\Juliana.png',
                'resp_rapida_cad_whats': '/Juliana',
                'resp_rapida_ja_cad_whats': '/ClienteJaCadastradoJu'
            },

            {
                'nome': 'MARI',
                'resp_rapida_cad_whats': '/Mari',
                'resp_rapida_ja_cad_whats': '/ClienteJaCadastradoMari'
            },

            {
                'nome': 'CRIS',
                'resp_rapida_cad_whats': '/Cris',
                'resp_rapida_ja_cad_whats': '/ClienteJaCadastradoCris'
            },

            {
                'nome': 'SIMONE',
                'resp_rapida_cad_whats': '/Simone1',
                'resp_rapida_ja_cad_whats': '/ClienteJaCadastradoSimo1'
            },
        ]

    def _obter_prox_nome(self, nome_atual):
        indice_nome_atual = self.lista_vendedoras.index(nome_atual)

        return self.lista_vendedoras[(indice_nome_atual + 1) % len(self.lista_vendedoras)]

    def _ler_ult_nome(self):
        with open(self.caminho_arq_log, 'r') as arq:
            return arq.read().strip()
            
    def _salvar_nome(self, nome):
        with open(self.caminho_arq_log, 'w') as arq:
            arq.write(nome)

    def _selecionar_vendedora(self):
        ult_nome = self._ler_ult_nome()

        prox_nome = self._obter_prox_nome(ult_nome)

        self._salvar_nome(prox_nome)

        return prox_nome

    def _obter_dados_vendedora(self, nome_vendedora):
        for vendedora in self.dic_vendedoras:
            if vendedora['nome'] == nome_vendedora:
                return vendedora


class Automacoes_manse:
    def __init__(self):
        self.planilhas = Planilhas()
        self.vendedoras = Vendedoras()
        pg.PAUSE = 0.5
        larg, alt = self._verif_tam_tela()
        self.regiao = (0, 0, round(larg/2), alt)
        print('')

    def _cad_cliente_manse(self, cnpj_cliente_original, nome_cliente, tel_cliente, num_inscricao, nome_vendedora):
        try:
            try:
            # Procurar o botão "Clientes" e clicar nele, se não o achar, clica no botão "Novo"
                btn_clientes = pg.locateCenterOnScreen(r'Imagens_manse\botao_clientes.png', region=self.regiao)
                pg.click(btn_clientes)
                sleep(0.2)
                btn_novo = pg.locateCenterOnScreen(r'Imagens_manse\botao_novo.png', region=self.regiao)
                pg.click(btn_novo)
            except:
                btn_novo = pg.locateCenterOnScreen(r'Imagens_manse\botao_novo.png', region=self.regiao)
                pg.click(btn_novo)

            # Clicar no botão "S"
            btn_s = pg.locateCenterOnScreen(r'Imagens_manse\botao_s_cad.png', region=self.regiao)
            pg.click(btn_s)

            # CLicar no botão "Manse cadastro"
            btn_manse_cad = pg.locateCenterOnScreen(r'Imagens_manse\botao_manse_cad.png', region=self.regiao)
            pg.click(btn_manse_cad)

            # Procurar o botão "Consultar"
            btn_consultar = pg.locateCenterOnScreen(r'Imagens_manse\botao_consultar.png', region=self.regiao)
            
            # Clicar na barra de pesquisa e colar o CNPJ
            pg.click(btn_consultar[0]-70, btn_consultar[1])

            # Copiar o CNPJ e colar na barra de pesquisa
            pyperclip.copy(cnpj_cliente_original)
            pg.hotkey('ctrl', 'v')

            # CLicar em "Consultar" e aguardar carregamento dos dados
            pg.click(btn_consultar)
            while True:
                try:
                    busca_dados_manse_rest = pg.locateCenterOnScreen(r'Imagens_manse\busca_dados_manse_rest.png', region=self.regiao)
                    if busca_dados_manse_rest:
                        sleep(0.1)
                except:
                    pyperclip.copy('')
                    break

            # Verificar se os dados foram preenchidos, senão, fecha a tela
            try:
                espaco_cnpj = pg.locateCenterOnScreen(r'Imagens_manse\espaco_cnpj.png', region=self.regiao)
                if espaco_cnpj:
                    btn_fechar_cad = pg.locateCenterOnScreen(r'Imagens_manse\botao_fechar_cad.png', region=self.regiao)
                    pg.click(btn_fechar_cad)
                    return False
            except:
                # Clicar em "Inserir" e depois apertar "enter"
                btn_inserir = pg.locateCenterOnScreen(r'Imagens_manse\botao_inserir.png', region=self.regiao)
                pg.click(btn_inserir)
                sleep(0.3)
                pg.press('enter')

            # Procurar "Insc. Estadual" e clicar nesse campo
            inscr_estadual = pg.locateCenterOnScreen(r'Imagens_manse\inscri_estad.png', region=self.regiao)
            pos_inscr_estadual = (inscr_estadual[0], inscr_estadual[1]+20)
            pg.click(pos_inscr_estadual)

            # Procurar "Suframa" e clicar nesse campo
            suframa = pg.locateCenterOnScreen(r'Imagens_manse\suframa.png', region=self.regiao)
            pg.click(suframa)
            sleep(0.3)

            # Verificar se há erro na Inscrição Estadual
            while True:
                try:
                    atencao_geral = pg.locateCenterOnScreen(r'Imagens_manse\atencao_geral.png', region=self.regiao)
                    if atencao_geral:
                        pg.press('enter')
                        pg.click(pos_inscr_estadual)
                        pg.click(pos_inscr_estadual)
                        pg.hotkey('ctrl', 'a')
                        pg.press('delete')
                        if num_inscricao:
                            pyperclip.copy(num_inscricao)
                            pg.hotkey('ctrl', 'v')
                            break
                        else:
                            sem_inscri_estad = pg.locateCenterOnScreen(r'Imagens_manse\aviso_para_inscri_estad.png', region=self.regiao)
                            pos_sem_inscri_estad = (sem_inscri_estad[0], sem_inscri_estad[1]+20)
                            pg.click(pos_sem_inscri_estad)
                            pg.click(pos_sem_inscri_estad[0]+20, pos_sem_inscri_estad[1])
                            pg.write('SEM INSCRICAO ESTADUAL')
                            break
                except:
                    pg.click(pos_inscr_estadual)
                    pg.hotkey('ctrl', 'a')
                    pg.hotkey('ctrl', 'c')
                    num_inscricao_manse = pyperclip.paste()
                    if num_inscricao_manse == '':
                        if num_inscricao:
                            pyperclip.copy(num_inscricao)
                            pg.hotkey('ctrl', 'v')
                            break
                        else:
                            sem_inscri_estad = pg.locateCenterOnScreen(r'Imagens_manse\aviso_para_inscri_estad.png', region=self.regiao)
                            pos_sem_inscri_estad = (sem_inscri_estad[0], sem_inscri_estad[1]+20)
                            pg.click(pos_sem_inscri_estad)
                            pg.click(pos_sem_inscri_estad[0]+20, pos_sem_inscri_estad[1])
                            pg.write('SEM INSCRICAO ESTADUAL')
                            break
                    else:
                        break
                    
            
            # Verificar e-mail


            # Procurar o campo "Celular" e colocar o número da cliente
            celular = pg.locateCenterOnScreen(r'Imagens_manse\celular.png', region=self.regiao)
            pg.click(celular)
            pyperclip.copy(tel_cliente)
            pg.hotkey('ctrl', 'v')

            # Procurar o campo "whatsapp" e escrever "paulo"
            campo_whats = pg.locateCenterOnScreen(r'Imagens_manse\whatsapp.png', region=self.regiao)
            pg.click(campo_whats)
            pg.write('paulo')
            pyperclip.copy('')

            # Verificar se o campo "Contato" tem o nome da cliente, senão, colocar o nome
            campo_contato = pg.locateCenterOnScreen(r'Imagens_manse\campo_contato_cad_manse.png', region=self.regiao)
            pos_contato = (campo_contato[0], campo_contato[1])
            pg.click(pos_contato[0], pos_contato[1]+20)
            pg.hotkey('ctrl', 'a')
            pg.hotkey('ctrl', 'c')
            nome_cliente_manse = pyperclip.paste()
            if nome_cliente_manse == '' and nome_cliente is not None:
                pyperclip.copy(nome_cliente)
                pg.hotkey('ctrl', 'v')
            else:
                pyperclip.copy('')

            # Selecionar vendedora e clicar em "OK"
            area_vendedora = pg.locateCenterOnScreen(r'Imagens_manse\area_vendedora.png', region=self.regiao)
            pg.click(area_vendedora[0], area_vendedora[1]+5)

            dados_vendedora = self.vendedoras._obter_dados_vendedora(nome_vendedora)

            cod_vendedora = pg.locateCenterOnScreen(dados_vendedora['img_cad_manse'], region=self.regiao)
            pg.click(cod_vendedora)

            btn_ok = pg.locateCenterOnScreen(r'Imagens_manse\ok_cad_vendedora.png', region=self.regiao)
            pg.click(btn_ok)

            # Clicar em no botão "Salvar" e depois em "enter"
            btn_salvar_manse = pg.locateCenterOnScreen(r'Imagens_manse\salvar_cad_manse.png', region=self.regiao)
            pg.click(btn_salvar_manse)
            sleep(0.4)
            pg.press('enter')

            print(' - CLIENTE CADASTRADO NO MANSE')

            return True, num_inscricao_manse if num_inscricao_manse else None
               
        except Exception as e:
            print(f'Erro em realizar ações no cadastro de clientes no manse:\n{e}')
            pg.screenshot('erro_screenshot.png')
            sys.exit(1)

    def _att_base_dados_mase(self):
        print(' - ATUALIZANDO PLANILHA "BASE CLIENTES MANSE"...')
        try:
            try:
            # Procurar o botão "Clientes" e clicar nele, se não o achar, clica no botão "Novo"
                btn_clientes = pg.locateCenterOnScreen(r'Imagens_manse\botao_clientes.png', region=self.regiao)
                pg.click(btn_clientes)
                sleep(0.2)
                btn_novo = pg.locateCenterOnScreen(r'Imagens_manse\botao_novo.png', region=self.regiao)
                pg.click(btn_novo)
            except:
                btn_novo = pg.locateCenterOnScreen(r'Imagens_manse\botao_novo.png', region=self.regiao)
                pg.click(btn_novo)

            # Procurar o botão "Procurar" e clicar nele
            btn_procurar = pg.locateCenterOnScreen(r'Imagens_manse\botao_procurar.png', region=self.regiao)
            pg.click(btn_procurar)

            # Procurar o botão "Listar"
            btn_listar = pg.locateCenterOnScreen(r'Imagens_manse\botao_listar.png', region=self.regiao)

            # Clicar na barra de pesquisa e deletar o que tiver lá
            pg.click(btn_listar[0]-100, btn_listar[1])
            pg.hotkey('ctrl', 'a')
            pg.press('delete')

            # Clicar no botão "Listar"
            pg.click(btn_listar)

            # Aguardar até que a informação de carregando suma
            while True:
                try:
                    info_carregando = pg.locateCenterOnScreen(r'Imagens_manse\info_carregando.png', region=self.regiao)
                    if info_carregando:
                        sleep(0.1)
                except:
                    break

            # Procurar o botão "Outros" e clicar nele
            btn_outros = pg.locateCenterOnScreen(r'Imagens_manse\botao_outros.png', region=self.regiao)
            pg.click(btn_outros)

            # Procurar o botão "Gerar txt (Excel)" e clicar nele
            btn_gerar_txt = pg.locateCenterOnScreen(r'Imagens_manse\botao_gerar_txt_excel.png', region=self.regiao)
            pg.click(btn_gerar_txt)

            # Aguardar até que a informação de gerando arquivo suma
            while True:
                try:
                    info_gerando_arq = pg.locateCenterOnScreen(r'Imagens_manse\info_gerando_arq.png', region=self.regiao)
                    if info_gerando_arq:
                        sleep(0.1)
                except:
                    sleep(0.3)
                    pg.press('enter')
                    break

            # Aguardar alguns segundos, procurar pelo botão "Fechar" e clicar nele
            sleep(0.3)
            btn_fechar = pg.locateCenterOnScreen(r'Imagens_manse\botao_fechar.png', region=self.regiao)
            pg.click(btn_fechar)
            sleep(0.5)

            # Atualizar planilha "Base de clientes manse"
            self.planilhas._att_plan_manse()
            print(' - PLANILHA ATUALIZADA COM SUCESSO!')

        except Exception as e:
            print(f'Erro em realizar ações na atualização da planilha do manse:\n{e}')
    
    def _verif_tam_tela(self):
        tam_tela = pg.size()

        larg = tam_tela[0]
        alt = tam_tela[1]

        return larg, alt


class Navegador:
    def __init__(self):
        self.automacoes = Automacoes_manse()
        self.api = Request_API()

        self.caminho_chrome = r'C:\Program Files\Google\Chrome\Application\chrome.exe'
        self.temp_dir_chrome = r'C:\ChromeDebug'
        self.porta_remota = 9222

        self.navegador, self.wait, self.acao = self._iniciar_navegador()
        self.abas = self.navegador.window_handles
        self._verif_carregamento_msgs()

    def _config_navegador(self):
        larg, alt = self.automacoes._verif_tam_tela()

        options = webdriver.ChromeOptions()
        
        argumentos = [
            '--block-new-web-contents',
            '--disable-notifications',
            '--no-default-browser-check',
            '--disable-features=ProtocolHandlers',
            '--lang=pt-BR',
            #'--headless',
            f'--window-position={larg/2},0',
            f'--window-size={larg/2},{alt}'
        ]

        for argumento in argumentos:
            options.add_argument(argumento)

        options.debugger_address = f'127.0.0.1:{self.porta_remota}'

        navegador = webdriver.Chrome(options=options)

        wait = WebDriverWait(navegador, timeout=10)

        acao = ActionChains(navegador)

        return navegador, wait, acao
    
    def _config_usar_nav_existente(self):
        comando_chrome = [
            self.caminho_chrome,
            f'--remote-debugging-port={self.porta_remota}',
            f'--user-data-dir={self.temp_dir_chrome}'
        ]

        subprocess.Popen(comando_chrome, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        sleep(2)

    def _iniciar_navegador(self):
        self._config_usar_nav_existente()

        print('ABRINDO NAVEGADOR...')

        navegador, wait, acao = self._config_navegador()

        navegador.get('https://web.whatsapp.com/')
        navegador.execute_script('window.open("https://contacts.google.com/", "_blank");')

        return navegador, wait, acao

    def _verif_carregamento_msgs(self):
        self.navegador.switch_to.window(self.abas[0])
        print('AGUARDANDO CARREGAMENTO DE MENSAGENS...')

        while True:
            try:
                carre_msgs = self.navegador.find_element(by=By.XPATH, value='//div[@class="x1c3i2sq x14ug900 xk82a7y x1sy10c2"]')
                if carre_msgs:
                    sleep(2)
                else:
                    break
            except:
                break

    def _cadastrar_cliente_celular(self, razao, nome_cliente, tel_cliente, cnpj_original):
        if nome_cliente:
            nome_contato = f'{razao} ({cnpj_original}) / {nome_cliente}'
        else:
            nome_contato = f'{razao} ({cnpj_original})'

        self.navegador.switch_to.window(self.abas[1])

        try:
            # Verificar se tem um contato sendo processado
            while True:
                try:
                    processando_contato = self.navegador.find_element(by=By.XPATH, value='/html/body/div[8]/aside/div/div').text
                    if processando_contato == 'Processando...':
                        sleep(1)
                    else:
                        break
                except:
                    break

            # Clicar em "Criar contato"
            btn_criar_contato = self.wait.until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[8]/c-wiz[1]/div/gm-coplanar-drawer/div/div/div/div[1]/div/div/div/div[1]/button')))
            btn_criar_contato.click()
            sleep(0.75)

            btn_criar_contato_menu = self.wait.until(EC.visibility_of_element_located((By.XPATH, '//li[@data-menu-item="CreateContact"]')))
            btn_criar_contato_menu.click()
            sleep(0.75)

            # Preenche o campo de "Nome"
            campo_nome = self.wait.until(EC.element_to_be_clickable((By.XPATH, '//input[@aria-label="Nome"]')))
            campo_nome.send_keys(nome_contato)
            sleep(0.75)

            # Preenche o campo de "Telefone"
            campo_tel = self.wait.until(EC.element_to_be_clickable((By.XPATH, '//input[@aria-label="Telefone"]')))
            campo_tel.send_keys(tel_cliente)
            sleep(0.75)

            # Clicar no botão "Salvar"
            btn_salvar = self.wait.until(EC.element_to_be_clickable((By.XPATH, '//button[@aria-label="Salvar"]')))
            btn_salvar.click()

            print(' - CLIENTE CADASTRADO NO CELULAR')
            sleep(0.75)
            
            return True

        except Exception as e:
            print(f'Erro em cadastrar no CELULAR:\n{e}')
            return False

    def _etiquetar_msg_whatsapp(self, nome_vendedora=None, tel_cliente=None, sem_praca=False):
        try:
            # Ir à aba do whatsapp
            self.navegador.switch_to.window(self.abas[0])

            # Clicar na aba de "Tudo"
            aba_tudo = self.wait.until(EC.element_to_be_clickable((By.XPATH, '//button[@id="all-filter"]')))
            aba_tudo.click()
            sleep(1)

            # Clicar na barra de pesquisa, deleta o que estiver lá e depois escreve o número do cliente
            barra_pesquisa = self.wait.until(EC.element_to_be_clickable((By.XPATH, '//div[@contenteditable="true"][@aria-label="Pesquisar"]')))
            barra_pesquisa.click()
            self.acao.key_down(Keys.CONTROL).send_keys('a').key_up(Keys.CONTROL).send_keys(Keys.DELETE).perform()
            barra_pesquisa.send_keys(tel_cliente)
            sleep(1)

            # Clicar com o botão direito na conversa
            conversa = self.wait.until(EC.element_to_be_clickable((By.XPATH, '//div[@role="gridcell"][@aria-colindex="2"]')))
            self.acao.context_click(conversa).perform()
            sleep(0.75)

            # Clicar em "Etiquetar conversa"
            etiquetar = self.wait.until(EC.element_to_be_clickable((By.XPATH, '//div[@aria-label="Etiquetar conversa"]')))
            etiquetar.click()
            sleep(0.75)

            # Escolhe a etiqueta com base no nome da vendedora ou, se for sem praça, seleciona "Sem Praça"
            if sem_praca:
                etiq_sem_praca = self.wait.until(EC.element_to_be_clickable((By.XPATH, '//span[@title="Sem Praça"]')))
                etiq_sem_praca.click()
            else:
                if nome_vendedora == "VIVIANE":
                    etiq_vivi = self.wait.until(EC.element_to_be_clickable((By.XPATH, '//span[@title="Viviane"]')))
                    etiq_vivi.click()

                elif nome_vendedora == "JULIANA":
                    etiq_ju = self.wait.until(EC.element_to_be_clickable((By.XPATH, '//span[@title="Juliana"]')))
                    etiq_ju.click()

                elif nome_vendedora == "SIMONE DJ":
                    etiq_simone_dj = self.wait.until(EC.element_to_be_clickable((By.XPATH, '//span[@title="Simone DJ"]')))
                    etiq_simone_dj.click()

                elif nome_vendedora == "LUCIANA":
                    etiq_luciane = self.wait.until(EC.element_to_be_clickable((By.XPATH, '//span[@title="Luciane"]')))
                    etiq_luciane.click()

                elif nome_vendedora == 'MARI':
                    menu_etiquetas = self.wait.until(EC.presence_of_element_located((By.XPATH, '//div[@class="x12lqup9 x1o1kx08 xubnuyq x1odjw0f"]')))
                    self.navegador.execute_script('arguments[0].scrollTop += 500;', menu_etiquetas)
                    etiq_mari = self.wait.until(EC.element_to_be_clickable((By.XPATH, '//span[@title="Mari"]')))
                    etiq_mari.click()

                elif nome_vendedora == 'CRIS':
                    menu_etiquetas = self.wait.until(EC.presence_of_element_located((By.XPATH, '//div[@class="x12lqup9 x1o1kx08 xubnuyq x1odjw0f"]')))
                    self.navegador.execute_script('arguments[0].scrollTop += 500;', menu_etiquetas)
                    etiq_mari = self.wait.until(EC.element_to_be_clickable((By.XPATH, '//span[@title="Cris"]')))
                    etiq_mari.click()

                elif nome_vendedora == 'SIMONE':
                    menu_etiquetas = self.wait.until(EC.presence_of_element_located((By.XPATH, '//div[@class="x12lqup9 x1o1kx08 xubnuyq x1odjw0f"]')))
                    self.navegador.execute_script('arguments[0].scrollTop += 500;', menu_etiquetas)
                    etiq_mari = self.wait.until(EC.element_to_be_clickable((By.XPATH, '//span[@title="Simone"]')))
                    etiq_mari.click()

            sleep(0.75)

            # Clicar em "Salvar"
            btn_salvar = self.wait.until(EC.element_to_be_clickable((By.XPATH, '//button[@class="x889kno x1a8lsjc xbbxn1n xxbr6pl x1n2onr6 x1rg5ohu xk50ysn x1f6kntn xyesn5m x1z11no5 xjy5m1g x1mnwbp6 x4pb5v6 x178xt8z xm81vs4 xso031l xy80clv x13fuv20 xu3j5b3 x1q0q8m5 x26u7qi x1v8p93f xogb00i x16stqrj x1ftr3km x1hl8ikr xfagghw x9dyr19 x9lcvmn xbtce8p x14v0smp xo8ufso xcjl5na x1k3x3db xuxw1ft xv52azi"]')))
            btn_salvar.click()

            print(' - MENSAGEM ETIQUETADA')
            sleep(0.75)

            return True

        except Exception as e:
            print(f'Erro em ETIQUETAR mensagem:\n{e}')
            return False

    def _enviar_msg_whatsapp(self, catalogo=None, dados_vendedora=None, tel_cliente=None, inscricao=None, sem_praca=False, ja_cad=False):
        try:
            # Ir à aba do whatsapp
            self.navegador.switch_to.window(self.abas[0])

            # Clicar na aba de "Tudo"
            aba_tudo = self.wait.until(EC.element_to_be_clickable((By.XPATH, '//button[@id="all-filter"]')))
            aba_tudo.click()
            sleep(1)

            # Clicar na barra de pesquisa, deleta o que estiver lá e depois escreve o número do cliente
            barra_pesquisa = self.wait.until(EC.element_to_be_clickable((By.XPATH, '//div[@contenteditable="true"][@aria-label="Pesquisar"]')))
            barra_pesquisa.click()
            self.acao.key_down(Keys.CONTROL).send_keys('a').key_up(Keys.CONTROL).send_keys(Keys.DELETE).perform()
            barra_pesquisa.send_keys(tel_cliente)
            sleep(1)

            # Clicar na conversa
            conversa = self.wait.until(EC.element_to_be_clickable((By.XPATH, '//div[@role="gridcell"][@aria-colindex="2"]')))
            conversa.click()
            sleep(1)

            # Clica na barra de "Digite uma mensagem" e envia o texto conforme a vendedora selecionada
            barra_digitar_msg = self.wait.until(EC.element_to_be_clickable((By.XPATH, '//div[@contenteditable="true"][@aria-label="Digite uma mensagem"]')))

            if sem_praca:
                barra_digitar_msg.send_keys('/SemPraça')
                sleep(1)
                enviar_com_texto = self.wait.until(EC.presence_of_element_located((By.XPATH, '//div[@contenteditable="true"][@aria-label="Digite uma mensagem"]')))
                enviar_com_texto.send_keys(Keys.ENTER)
                sleep(0.75)
                enviar_com_texto.send_keys(Keys.ENTER)
                sleep(1)
            else:
                if ja_cad:
                    barra_digitar_msg.send_keys(dados_vendedora['resp_rapida_ja_cad_whats'])
                    sleep(1)
                    barra_digitar_msg.send_keys(Keys.ENTER)
                else:
                    barra_digitar_msg.send_keys(dados_vendedora['resp_rapida_cad_whats'])
                    sleep(1)
                    barra_digitar_msg.send_keys(Keys.ENTER)

                # Clicar no "+" e enviar o catálogo
                mais_whats = self.wait.until(EC.element_to_be_clickable((By.XPATH, '//span[@data-icon="plus"]')))
                mais_whats.click()
                sleep(1)
                input_mais_whats = self.wait.until(EC.presence_of_element_located((By.XPATH, '//input[@accept="*"]')))
                input_mais_whats.send_keys(catalogo)
                sleep(3)

                # Clicar em "Enviar"
                enviar = self.wait.until(EC.element_to_be_clickable((By.XPATH, '//div[@role="button"][@aria-label="Enviar"]')))
                enviar.click()
                sleep(1)

                # Verifica se está sem inscricão estadual e envia a mensagem
                if not inscricao:
                    barra_digitar_msg.send_keys('/InscriçãoEstadual')
                    sleep(1)
                    enviar_com_texto = self.wait.until(EC.presence_of_element_located((By.XPATH, '//div[@contenteditable="true"][@aria-label="Digite uma mensagem"]')))
                    enviar_com_texto.send_keys(Keys.ENTER)
                    sleep(0.75)
                    enviar_com_texto.send_keys(Keys.ENTER)
                    sleep(1)
            
            print(' - MENSAGEM ENVIADA')
            sleep(0.75)

            return True

        except Exception as e:
            print(f'Erro em ENVIAR mensagem:\n{e}')
            return False


class Execucao_geral:
    def __init__(self):
        self.api = Request_API()
        self.planilhas = Planilhas()
        self.formatacoes = Formatacoes()
        self.vendedoras = Vendedoras()
        self.automacoes_manse = Automacoes_manse()
        self.navegador = Navegador()
        self.caminho_catalogo = r'c:\Users\Cliente\OneDrive\Desktop\Verificar Cadastros\TALGUI - INVERNO 2025.pdf'
    
    def _execucao_geral(self):
        plan_clientes = self.planilhas._carregar_planilha_clientes()

        contador = 0
        tempo_inicio = 0

        for linha in plan_clientes.iter_rows(min_row=2):
            nome_cliente = linha[1].value
            tel_cliente = linha[2].value
            cnpj_cliente = str(linha[3].value)
            
            # Verifica se a célula do CNPJ está vazia, caso esteja, finaliza o processo
            if not linha[3].value:
                break

            if linha[4].value in ['NEGADO', 'ERRO AO CONSULTAR O CNPJ', 'VERIFICAR VENDEDORA E HISTÓRICO DE COMPRA','ERRO EM CADASTRAR NO MANSE', 'JULIANA', 'VIVIANE', 'MARI', 'SIMONE', 'CRIS', 'LUCIANE', 'LUCIANA', 'SIMONE DJ']:
                continue

            cnpj_consulta_cliente = self.formatacoes._formatar_cnpj_consulta(cnpj_cliente)

            resp = self.api._consultar_cnpj(cnpj_consulta_cliente)

            contador, tempo_inicio = self.api._contar_execucoes(contador, tempo_inicio)

            if not resp:
                linha[4].value = "ERRO AO CONSULTAR O CNPJ"
                
                self.planilhas._salvar_planilha(self.planilhas.planilha_clientes, self.planilhas.caminho_plan_clientes)
                continue

            print(f'VERIFICANDO CNPJ: {cnpj_cliente}')

            status, dt_cadastro, nome_vendedora = self._verif_aprovacao_cadastro(resp, cnpj_consulta_cliente)

            self._verif_status_executar_acoes(
                resp,
                linha, 
                status, 
                dt_cadastro,
                nome_vendedora, 
                nome_cliente,
                tel_cliente,
                cnpj_consulta_cliente
            )

            self.planilhas._salvar_planilha(self.planilhas.planilha_clientes, self.planilhas.caminho_plan_clientes)

    def _verif_aprovacao_cadastro(self, resp, cnpj_consulta_cliente):
        situacao_se_ativa, se_ramo = self.api._verif_situacao_ramo(resp)

        if situacao_se_ativa and se_ramo:
            bairro, cidade, uf = self.api._extrair_endereco(resp)

            praca = self.planilhas._verif_praca(cnpj_consulta_cliente, bairro, cidade, uf)
            dt_cadastro, nome_vendedora, cadastrado = self.planilhas._verif_ja_cadastrado(cnpj_consulta_cliente)

            if praca:
                if cadastrado:
                    status = 'JÁ CADASTRADO - COM PRAÇA'
                else:
                    status = 'FAZER CADASTRO - COM PRAÇA'
            else:
                status = 'REPROVADO - SEM PRAÇA'

        else:
            if not situacao_se_ativa:
                status = 'REPROVADO - CNPJ INAPTO'
                
            if not se_ramo:
                status = 'REPROVADO - NÃO RAMO'
            
            dt_cadastro, nome_vendedora = None, None

        return status, dt_cadastro, nome_vendedora

    def _verif_status_executar_acoes(self, resp, linha, status, dt_cadastro, nome_vendedora, nome_cliente, tel_cliente, cnpj_consulta_cliente):
        razao = self.api._verif_razao(resp)
        num_inscricao = self.api._verif_inscricao(resp)
        cnpj_cliente_original = self.formatacoes._formatar_cnpj_original(cnpj_consulta_cliente)

        if status == 'FAZER CADASTRO - COM PRAÇA':
            # Selecionar vendedora e pegar seus dados
            nome_vendedora_log = self.vendedoras._selecionar_vendedora()
            dados_vendedora_log = self.vendedoras._obter_dados_vendedora(nome_vendedora_log)

            # Cadastrar cliente no manse
            retorno_cad_manse = self.automacoes_manse._cad_cliente_manse(cnpj_cliente_original, nome_cliente, tel_cliente, num_inscricao, nome_vendedora_log)

            if retorno_cad_manse:
                # Colocar data do cadastro na planilha
                linha[0].value =  datetime.now().strftime('%d/%m/%Y')

                # Colocar o nome da vendedora e o status da ação na célula
                linha[4].value = nome_vendedora_log
                linha[5].value = 'CADASTRADO NO MANSE'

                # Cadastrar cliente no celular
                cad_celular = self.navegador._cadastrar_cliente_celular(razao, nome_cliente, tel_cliente, cnpj_cliente_original)
                if cad_celular:
                    linha[6].value = 'SIM'
                else:
                    linha[6].value = 'ERRO'

                # Etiquetar mensagem no whatsapp
                etiq_msg_whats = self.navegador._etiquetar_msg_whatsapp(nome_vendedora_log, tel_cliente)
                if etiq_msg_whats:
                    linha[7].value = 'SIM'
                else:
                    linha[7].value = 'ERRO'

                # Enviar mensagem à cliente
                enviar_msg = self.navegador._enviar_msg_whatsapp(self.caminho_catalogo, dados_vendedora_log, tel_cliente, num_inscricao)
                if enviar_msg:
                    linha[8].value = 'SIM'
                else:
                    linha[8].value = 'ERRO'

                # Verifica se os passos anteriores deram certo, caso sim, atualiza a validação
                linha[9].value = f'REPASSAR À VENDEDORA {nome_vendedora_log}' if cad_celular and etiq_msg_whats and enviar_msg else None

                # Atualizar planilha "Base de dados manse"
                self.automacoes_manse._att_base_dados_mase()
            else:
                linha[4].value = 'ERRO EM CADASTRAR NO MANSE'

        elif status == 'JÁ CADASTRADO - COM PRAÇA':
            linha[0].value = dt_cadastro

            linha[4].value = nome_vendedora if nome_vendedora else 'VERIFICAR VENDEDORA E HISTÓRICO DE COMPRA'
            linha[5].value = 'CLIENTE JÁ CADASTRADO'

            # Cadastrar cliente no celular
            cad_celular = self.navegador._cadastrar_cliente_celular(razao, nome_cliente, tel_cliente, cnpj_cliente_original)
            if cad_celular:
                linha[6].value = 'SIM'
            else:
                linha[6].value = 'ERRO'

            # Etiquetar mensagem no whatsapp e enviar mensagem se tiver vendedora
            if nome_vendedora:
                dados_vendedora_log = self.vendedoras._obter_dados_vendedora(nome_vendedora)

                etiq_msg_whats = self.navegador._etiquetar_msg_whatsapp(nome_vendedora, tel_cliente)
                if etiq_msg_whats:
                    linha[7].value = 'SIM'
                else:
                    linha[7].value = 'ERRO'

                # Enviar mensagem à cliente
                enviar_msg = self.navegador._enviar_msg_whatsapp(self.caminho_catalogo, dados_vendedora_log, tel_cliente, num_inscricao, ja_cad=True)
                if enviar_msg:
                    linha[8].value = 'SIM'
                else:
                    linha[8].value = 'ERRO'
            
                # Verifica se os passos anteriores deram certo, caso sim, atualiza a validação
                linha[9].value = f'REPASSAR À VENDEDORA {nome_vendedora}' if cad_celular and etiq_msg_whats and enviar_msg else None
            else:
                linha[7].value = 'NÃO'
                linha[8].value = 'NÃO'
                linha[9].value = 'VERIFICAR VENDEDORA E HISTÓRICO DE COMPRA'

        elif status == 'REPROVADO - SEM PRAÇA':
            linha[0].value =  datetime.now().strftime('%d/%m/%Y')
            linha[4].value = 'NEGADO'
            linha[5].value = 'SEM PRAÇA'
            linha[6].value = 'NÃO'

            # Etiquetar mensagem no whatsapp
            etiq_msg_whats = self.navegador._etiquetar_msg_whatsapp(tel_cliente=tel_cliente, sem_praca=True)
            if etiq_msg_whats:
                linha[7].value = 'SIM'
            else:
                linha[7].value = 'ERRO'

            # Enviar mensagem à cliente
            enviar_msg = self.navegador._enviar_msg_whatsapp(tel_cliente=tel_cliente, sem_praca=True)
            if enviar_msg:
                linha[8].value = 'SIM'
            else:
                linha[8].value = 'ERRO'

            # Validação de dados
            linha[9].value = 'SEM PRAÇA - COLOCAR NA PLANILHA ONLINE'

            # Colocar razão e inscrição nas células finais
            linha[10].value = razao
            linha[11].value = num_inscricao if num_inscricao else 'NÃO TEM'

        elif status == 'REPROVADO - CNPJ INAPTO':
            print(' - CNPJ INAPTO')
            linha[0].value =  datetime.now().strftime('%d/%m/%Y')
            linha[4].value = 'NEGADO'
            linha[5].value = 'CNPJ INAPTO'
            linha[6].value = 'NÃO'
            linha[7].value = 'NÃO'
            linha[8].value = 'NÃO'
            linha[9].value = 'CNPJ INAPTO - VERIFICAR CNPJ'

            # Colocar razão e inscrição nas células finais
            linha[10].value = razao
            linha[11].value = num_inscricao if num_inscricao else 'NÃO TEM'

        elif status == 'REPROVADO - NÃO RAMO':
            print(' - CNPJ NÃO É DO RAMO DE VESTUÁRIO')
            linha[0].value =  datetime.now().strftime('%d/%m/%Y')
            linha[4].value = 'NEGADO'
            linha[5].value = 'CNPJ NÃO É DO RAMO DE VESTUÁRIO'
            linha[6].value = 'NÃO'
            linha[7].value = 'NÃO'
            linha[8].value = 'NÃO'
            linha[9].value = 'CNPJ NÃO É DO RAMO DE VESTUÁRIO - VERIFICAR CNPJ'

            # Colocar razão e inscrição nas células finais
            linha[10].value = razao
            linha[11].value = num_inscricao if num_inscricao else 'NÃO TEM'


if __name__ == '__main__':
    print('='*60)
    print('INICIANDO PROCESSO!'.center(60))
    print('='*60)

    app = Execucao_geral()
    app._execucao_geral()

    # 07.526.557/0116-59
    # 34.884.173/0001-31
    
    print('='*60)
    print('PROCESSO FINALIZADO!'.center(60))
    print('='*60)
